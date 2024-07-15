import streamlit as st
from streamlit_lottie import st_lottie
import requests
import pandas as pd
import openpyxl
import tempfile
import base64
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.drawing.image import Image
from langchain_openai import ChatOpenAI
from dotenv import load_dotenv
from io import BytesIO
import openai
import plotly.express as px
import re
from PIL import Image
import altair as alt
from PIL import Image as PILImage
from openpyxl.drawing.image import Image as OpenpyxlImage
import numpy as np
import random
import os
import json
import pickle
from PIL import Image
from nltk.stem import WordNetLemmatizer



headers = {
    "authorization": os.getenv("OPENAI_API_KEY"),
    "content-type": "application/json"
}
# Constants
LOTTIE_URL = 'https://lottie.host/a65c288c-0fca-43eb-a6db-96f641014e25/Kw3wnMn3vd.json'
PAGE_TITLE = "Yazaki Finance Assistant"
PAGE_ICON = ":bar_chart:"
WELCOME_VIDEO_URL = "https://youtu.be/AdYNFipk8dE?si=tibMqjBhhLhnFYMF"
IMAGE_PATH = "1ksp_yazaki.png"

def load_lottie_url(url: str) -> dict:
    """
    Load a Lottie animation from a given URL.
    
    Args:
    - url (str): URL to the Lottie JSON file.
    
    Returns:
    - dict: The Lottie animation JSON if successful, None otherwise.
    """
    try:
        response = requests.get(url)
        response.raise_for_status()
        return response.json()
    except requests.exceptions.RequestException as error:
        st.error(f"Error loading Lottie animation: {error}")
        return None
def is_numeric(value: str) -> bool:
    return value.isdigit()

def add_image(sheet, image_path):
    # Load the image using PIL
    pil_img = PILImage.open(image_path)
    # Resize the image
    resized_img = pil_img.resize((170, 35))
    # Save the resized image to a BytesIO object
    img_byte_arr = BytesIO()
    resized_img.save(img_byte_arr, format='PNG')
    # Convert the BytesIO object to an Openpyxl image
    img_byte_arr.seek(0)
    openpyxl_img = OpenpyxlImage(img_byte_arr)
    # Add the image to the sheet
    sheet.add_image(openpyxl_img, 'A1')

def process_excel(file_path: str):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook["BG affectée"]
    
    row_count = sheet.max_row
    column_count = sheet.max_column
    target_col = 19
    first_data_row = 5

    # Count the number of "Solde" columns in the header row (row 4)
    solde_count = 0
    solde_columns = []
    for col in range(1, column_count + 1):
        cell_value = sheet.cell(row=4, column=col).value
        if cell_value and isinstance(cell_value, str) and "Solde" in cell_value:
            solde_count += 1
            solde_columns.append(col)
    
    # Check if there are exactly 9 "Solde" columns
    if solde_count == 9:
        # Delete the oldest year column (the first column with "Solde" in the header)
        first_year_col = solde_columns[0]
        sheet.delete_cols(first_year_col)
    
    # Find the last column with "Solde" in the header row 4
    last_col = column_count - 1
    while last_col > 1:
        cell_value = sheet.cell(row=4, column=last_col+1).value
        if cell_value is not None and isinstance(cell_value, str) and "Solde" in cell_value:
            break
        last_col -= 1

    # Get the header text
    header_text = sheet.cell(row=4, column=last_col+1).value
    if header_text and isinstance(header_text, str) and is_numeric(header_text[-4:]):
        current_year = int(header_text[-4:])
    else:
        st.error(f"The header in column {last_col + 1} does not contain a valid year: {header_text}")
        return None

    # Determine the next year
    next_year = current_year + 1

    # Insert a new column at position S (19th column)
    sheet.insert_cols(target_col)

    # Add a new column for the next year in column S
    sheet.cell(row=4, column=target_col).value = f"Solde 31.03.{next_year}"

    # Find the last data row based on column D
    last_data_row = sheet.max_row

    # Apply the formula to the new column from the first data row to the last data row using a loop
    for row in range(first_data_row, last_data_row + 1):
        formula = f"=SUMIF('BG SAP'!E:E, 'BG affectée'!D{row}, 'BG SAP'!K:K)"
        try:
            sheet.cell(row=row, column=target_col).value = formula
        except Exception as error:
            st.error(f"Error applying formula to cell {row}, {target_col}: {error}")
    
    
    
    # Set specific cells to 0
    zero_cells = ['S321', 'S48', 'S142', 'S143', 'S22','S35','S23']
    for cell in zero_cells:
        sheet[cell].value = 0
    
    # Set number format for column S
    for row in range(1, sheet.max_row + 1):
        sheet[f"S{row}"].number_format = '_(* #,##0_);_(* (#,##0);_(* "-"??_);_(@_)'
    
    return workbook



def bilan_actif(workbook):
    image_path = r'1ksp_yazaki.png'
    # Add a new worksheet with name "Actif"
    actif = workbook.create_sheet("Actif")
    add_image(actif, image_path)
    # Get years from the BG affectée sheet
    bg_affectee = workbook["BG affectée"]
    year1 = bg_affectee["S4"].value
    year2 = bg_affectee["R4"].value  
    
    # Set values in specific cells
    cell_values = {
        "A3": "YAZAKI AUTOMOTIVE PRODUCTS TUNISIA Sarl",
        "B6": "BILAN ARRETE au 31 mars 2023",
        "B7": "(exprimé en Dinars Tunisiens)",
        "B9": "ACTIFS",
        "C9": "Notes",
        "B10": "ACTIFS NON COURANTS",
        "B12": "Immobilisations incorporelles",
        "B13": "Moins: Amortissements II",
        "B14": "Total des immobilisations incorporelles",
        "B16": "Immobilisations corporelles",
        "B17": "Moins: Amortissements IC",
        "B18": "Total des immobilisations corporelles",
        "B20": "Immobilisations financières",
        "B21": "Moins: provisions IF",
        "B22": "Net",
        "B24": "Autres actifs non courants",
        "B26": "Total des actifs non courants",
        "B28": "Stocks",
        "B29": "Moins: Provisions pour dépréciation Stock",
        "B30": "Net",
        "B32": "Clients et comptes rattachés",
        "B33": "Moins : Provisions sur clients",
        "B34": "Net",
        "B36": "Autres actifs courants",
        "B37": "Moins: Provisions pour dépréciation Actifs",
        "B38": "Net",
        "B40": "Liquidités et équivalents de liquidités",
        "C40": "III.6",
        "B42": "Total des actifs courants",
        "B44": "TOTAL DES ACTIFS",
        "C14": "III.1",
        "C18": "III.1",
        "C22": "III.2",
        "C28": "III.3",
        "C32": "III.4",
        "C36": "III.5",
        "D9": f"Solde au 31 mars {year1}",
        "E9": f"Solde au 31 mars {year2}"
    }
    
    for cell, value in cell_values.items():
        actif[cell] = value

    # Set formulas in specific cells only if the corresponding cell in column B is not empty
    for row in range(12, 45):
        if actif[f"B{row}"].value:
            actif[f"D{row}"] = f"=SUMIF('BG affectée'!U:U,Actif!B{row},'BG affectée'!S:S)"
            actif[f"E{row}"] = f"=SUMIF('BG affectée'!U:U,Actif!B{row},'BG affectée'!R:R)"

    # Apply the necessary summary formulas
    summary_formulas = {
        "D14": "=SUM(D12:D13)",
        "E14": "=SUM(E12:E13)",
        "D18": "=SUM(D16:D17)",
        "E18": "=SUM(E16:E17)",
        "D22": "=SUM(D20:D21)",
        "E22": "=SUM(E20:E21)",
        "D26": "=SUM(D22,D18,D14)",
        "E26": "=SUM(E22,E18,E14)",
        "D30": "=SUM(D28:D29)",
        "E30": "=SUM(E28:E29)",
        "D34": "=SUM(D32:D33)",
        "E34": "=SUM(E32:E33)",
        "D38": "=SUM(D36:D37)",
        "E38": "=SUM(E36:E37)",
        "D42": "=SUM(D40,D38,D34,D30)",
        "E42": "=SUM(E40,E38,E34,E30)",
        "D44": "=SUM(D42,D26)",
        "E44": "=SUM(E42,E26)"
    }
    
    for cell, formula in summary_formulas.items():
        actif[cell] = formula
        actif[cell].font = Font(name="Times New Roman", bold=True)

    # Apply font properties to all cells
    font_times_new_roman = Font(name="Times New Roman", italic=False)
    for row in actif.iter_rows():
        for cell in row:
            cell.font = font_times_new_roman

    # Bold specific cells
    bold_cells = [
        "B14", "B22", "B18", "B6", "B9", "B10", "B15", "B19", "B22", "B26",
        "B30", "B38", "B42", "B44", "B34", "C9", "C14", "C18", "C22", "C28",
        "C32", "C36", "C40", "D44", "E44", "D42", "E42", "D38", "E38", "D34",
        "E34", "D30", "E30", "D26", "E26", "D22", "E22", "D18", "E18", "D14",
        "E14", "D9", "E9"
    ]
    for cell in bold_cells:
        actif[cell].font = Font(name="Times New Roman", bold=True)

    # Set column widths
    actif.column_dimensions['B'].width = 40
    actif.column_dimensions['C'].width = 10
    actif.column_dimensions['D'].width = 20
    actif.column_dimensions['E'].width = 20

    # Set alignment properties
    center_alignment = Alignment(horizontal="center", vertical="center")
    for cell in ["D9", "E9", "B6", "B7"]:
        actif[cell].alignment = center_alignment

    actif.merge_cells("B6:E6")
    actif.merge_cells("B7:E7")

    # Wrap text in headers
    for cell in ["D9", "E9"]:
        actif[cell].alignment = Alignment(wrap_text=True, horizontal="center")

    # Set gridlines visibility
    actif.sheet_view.showGridLines = False

    # Set number format for columns D and E to accounting format without symbol
    for col in ['D', 'E']:
        for row in range(1, actif.max_row + 1):
            actif[f"{col}{row}"].number_format = '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'

    return workbook

def bilan_actif_ang(workbook):
    image_path = r'1ksp_yazaki.png'
    # Add a new worksheet with name "Actif"
    actif = workbook.create_sheet("Actif ANG")
    add_image(actif, image_path)
    # Get years from the BG affectée sheet
    bg_affectee = workbook["BG affectée"]
    year1 = bg_affectee["S4"].value
    year2 = bg_affectee["R4"].value  
    
    # Set values in specific cells
    cell_values = {
        "A3": "YAZAKI AUTOMOTIVE PRODUCTS TUNISIA Sarl",
        "B6": f"BALANCE SHEET AS OF  March 31st,{year1}",
        "B7": "(in Tunisian Dinars)",
        "B9": "ASSETS",
        "C9": "Notes",
        "B10": "NON-CURRENT ASSETS",
        "B12": "Intangible Assets",
        "B13": "Depreciation II",
        "B14": "Total of  Intangible Assets",
        "B16": "Tangible Assets",
        "B17": "Depreciation IC",
        "B18": "Total of  Tangible Assets",
        "B20": "Financial Assets",
        "B21": "Provisions IF",
        "B22": "Net",
        "B24": "Other non-current assets",
        "B26": "Total Non-current assets",
        "B28": "Inventories",
        "B29": "Provisions for depreciation of inventories",
        "B30": "Net",
        "B32": "Trade and other receivables",
        "B33": "Provisions for receivables",
        "B34": "Net",
        "B36": "Other current assets",
        "B37": "Provisions for impairment of assets ",
        "B38": "Net",
        "B40": "Cash & cash equivalents",
        "C40": "III.6",
        "B42": "Total Current assets",
        "B44": "TOTAL ASSETS",
        "C14": "III.1",
        "C18": "III.1",
        "C22": "III.2",
        "C28": "III.3",
        "C32": "III.4",
        "C36": "III.5",
        "D9": f"Balance as of  March 31st, {year1}",
        "E9": f"Balance as of  March 31st, {year2}"
    }
    
    for cell, value in cell_values.items():
        actif[cell] = value

    # Set formulas in specific cells only if the corresponding cell in column B is not empty
    for row in range(12, 45):
        if actif[f"B{row}"].value:
            actif[f"D{row}"] = f"=+'Actif'!D{row}"
            actif[f"E{row}"] = f"=+'Actif'!E{row}"

    # Apply the necessary summary formulas
    summary_formulas = {
        "D14": "=SUM(D12:D13)",
        "E14": "=SUM(E12:E13)",
        "D18": "=SUM(D16:D17)",
        "E18": "=SUM(E16:E17)",
        "D22": "=SUM(D20:D21)",
        "E22": "=SUM(E20:E21)",
        "D26": "=SUM(D22,D18,D14)",
        "E26": "=SUM(E22,E18,E14)",
        "D30": "=SUM(D28:D29)",
        "E30": "=SUM(E28:E29)",
        "D34": "=SUM(D32:D33)",
        "E34": "=SUM(E32:E33)",
        "D38": "=SUM(D36:D37)",
        "E38": "=SUM(E36:E37)",
        "D42": "=SUM(D40,D38,D34,D30)",
        "E42": "=SUM(E40,E38,E34,E30)",
        "D44": "=SUM(D42,D26)",
        "E44": "=SUM(E42,E26)"
    }
    
    for cell, formula in summary_formulas.items():
        actif[cell] = formula
        actif[cell].font = Font(name="Times New Roman", bold=True)

    # Apply font properties to all cells
    font_times_new_roman = Font(name="Times New Roman", italic=False)
    for row in actif.iter_rows():
        for cell in row:
            cell.font = font_times_new_roman

    # Bold specific cells
    bold_cells = [
        "B14", "B22", "B18", "B6", "B9", "B10", "B15", "B19", "B22", "B26",
        "B30", "B38", "B42", "B44", "B34", "C9", "C14", "C18", "C22", "C28",
        "C32", "C36", "C40", "D44", "E44", "D42", "E42", "D38", "E38", "D34",
        "E34", "D30", "E30", "D26", "E26", "D22", "E22", "D18", "E18", "D14",
        "E14", "D9", "E9"
    ]
    for cell in bold_cells:
        actif[cell].font = Font(name="Times New Roman", bold=True)

    # Set column widths
    actif.column_dimensions['B'].width = 40
    actif.column_dimensions['C'].width = 10
    actif.column_dimensions['D'].width = 20
    actif.column_dimensions['E'].width = 20

    # Set alignment properties
    center_alignment = Alignment(horizontal="center", vertical="center")
    for cell in ["D9", "E9", "B6", "B7"]:
        actif[cell].alignment = center_alignment

    actif.merge_cells("B6:E6")
    actif.merge_cells("B7:E7")

    # Wrap text in headers
    for cell in ["D9", "E9"]:
        actif[cell].alignment = Alignment(wrap_text=True, horizontal="center")

    # Set gridlines visibility
    actif.sheet_view.showGridLines = False

    # Set number format for columns D and E to accounting format without symbol
    for col in ['D', 'E']:
        for row in range(1, actif.max_row + 1):
            actif[f"{col}{row}"].number_format = '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'

    return workbook

    # Add a new worksheet with name "P&L"
    pnl = workbook.create_sheet("P&L")
    
    # Get years from the BG affectée sheet
    bg_affectee = workbook["BG affectée"]
    year1 = bg_affectee["S4"].value
    year2 = bg_affectee["R4"].value 
    year3 = bg_affectee["Q4"].value 
    
    pnl["A3"] = "YAZAKI AUTOMOTIVE PRODUCTS TUNISIA Sarl"
    pnl["B6"] = f"ETAT DE RESULTAT ARRETE au 31 mars {year1}"
    pnl["B7"] = "(exprimé en Dinars Tunisiens)"
    pnl["B8"] = "Résultat"
    pnl["B10"] = "PRODUITS D'EXPLOITATION"
    pnl["B12"] = "Revenus"
    pnl["C12"] = "IV.1"
    pnl["B13"] = "Autres produits d'exploitation"
    pnl["C13"] = "IV.2"
    pnl["B15"] = "Total des produits d'exploitation"
    pnl["B17"] = "CHARGES D'EXPLOITATION"
    pnl["B19"] = "Variation de stocks de PF et encours"
    pnl["C19"] = "IV.3"
    pnl["B20"] = "Achats d'approvisionnements consommés"
    pnl["C20"] = "IV.3"
    pnl["B21"] = "Charges de personnel"
    pnl["C21"] = "IV.4"
    pnl["B22"] = "Dotations aux amortissements et aux provisions "
    pnl["C22"] = "IV.5"
    pnl["B23"] = "Autres Charges d'exploitation"
    pnl["C23"] = "IV.6"
    pnl["B25"] = "Total des charges d'exploitation"
    pnl["B27"] = "RESULTAT D'EXPLOITATION"
    pnl["B29"] = "Charges financières nettes"
    pnl["C29"] = "IV.7"
    pnl["B30"] = "Autres gains ordinaires"
    pnl["C30"] = "IV.8"
    pnl["B31"] = "Autres pertes ordinaires"
    pnl["C31"] = "IV.9"
    pnl["B33"] = "RESULTAT DES ACTIVITES ORDINAIRES AVANT IMPOT"
    pnl["B35"] = "Impôts sur les bénéfices"
    pnl["C35"] = "IV.10"
    pnl["B37"] = "RESULTAT NET DE L'EXERCICE"
    
    # Set headers for the years
    pnl["D8"] = f"Du 1er Avril {year2} au 31 mars {year1}"
    pnl["E8"] = f"Du 1er Avril {year3} au 31 mars {year2}"
    
    # Apply font properties to all cells
    font_times_new_roman = Font(name="Times New Roman", italic=False)
    for row in pnl.iter_rows():
        for cell in row:
            cell.font = font_times_new_roman
    
    # Bold specific cells
    bold_cells = [
        "A3", "B6", "B8", "B10", "B15", "B17", "B25", "B27", "B33", "B37",
        "D8", "E8", "B12", "B13", "B19", "B20", "B21", "B22", "B23", "B29", "B30", "B31", "B35",
        "C11","C12","C18","C19","C20","C21","C22","C23","C29","C30","C31","C35"
    ]
    for cell in bold_cells:
        pnl[cell].font = Font(name="Times New Roman", bold=True)
    
    # Set column widths
    pnl.column_dimensions['B'].width = 40
    pnl.column_dimensions['C'].width = 10
    pnl.column_dimensions['D'].width = 20
    pnl.column_dimensions['E'].width = 20
    
    # Set formulas in specific cells only if the corresponding cell in column B is not empty
    rows_for_negative_formula = [12, 13, 29, 30, 31, 35]
    rows_for_positive_formula = [19, 20, 21, 22, 23]
    
    for row in rows_for_negative_formula:
        if pnl[f"B{row}"].value:
            pnl[f"D{row}"] = f"=-SUMIF('BG affectée'!U:U,'P&L'!B{row},'BG affectée'!S:S)"
            pnl[f"E{row}"] = f"=-SUMIF('BG affectée'!U:U,'P&L'!B{row},'BG affectée'!R:R)"
    
    for row in rows_for_positive_formula:
        if pnl[f"B{row}"].value:
            pnl[f"D{row}"] = f"=SUMIF('BG affectée'!U:U,'P&L'!B{row},'BG affectée'!S:S)"
            pnl[f"E{row}"] = f"=SUMIF('BG affectée'!U:U,'P&L'!B{row},'BG affectée'!R:R)"
    
    # Apply the necessary summary formulas
    summary_formulas = {
        "D15": "=SUM(D12:D13)",
        "E15": "=SUM(E12:E13)",
        "D25": "=SUM(D19:D23)",
        "E25": "=SUM(E19:E23)",
        "D27": "=D15-D25",
        "E27": "=E15-E25",
        "D33": "=SUM(D27:D31)",
        "E33": "=SUM(E27:E31)",
        "D37": "=D33+D35",
        "E37": "=E33+E35"
    }
    
    for cell, formula in summary_formulas.items():
        pnl[cell] = formula
        pnl[cell].font = Font(name="Times New Roman", bold=True)
            
    # Set alignment properties
    center_alignment = Alignment(horizontal="center", vertical="center")
    pnl["D9"].alignment = center_alignment
    pnl["E9"].alignment = center_alignment
    pnl["B6"].alignment = center_alignment
    pnl["B7"].alignment = center_alignment
    pnl.merge_cells("B6:E6")
    pnl.merge_cells("B7:E7")

    # Hide all columns first
    for col in range(1, pnl.max_column + 1):
        pnl.column_dimensions[openpyxl.utils.get_column_letter(col)].hidden = True

    # Unhide specific columns
    columns_to_keep = ['A','B', 'C', 'D', 'E']
    for col in columns_to_keep:
        pnl.column_dimensions[col].hidden = False

    # Hide all rows first
    for row in range(1, pnl.max_row + 1):
        pnl.row_dimensions[row].hidden = True

    # Unhide specific rows
    rows_to_keep = range(1, 38)  # Modify this range based on your specific requirements
    for row in rows_to_keep:
        pnl.row_dimensions[row].hidden = False

    # Wrap text
    pnl["D8"].alignment = Alignment(wrap_text=True, horizontal="center")
    pnl["E8"].alignment = Alignment(wrap_text=True, horizontal="center")
    pnl["B33"].alignment = Alignment(wrap_text=True, horizontal="center")

    # Set gridlines visibility
    pnl.sheet_view.showGridLines = False
    
    # Set number format for columns D and E to accounting format without symbol
    for col in ['D', 'E']:
        for row in range(1, pnl.max_row + 1):
            pnl[f"{col}{row}"].number_format = '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'

    return workbook
    
def bilan_passif(workbook):
    image_path = r'1ksp_yazaki.png'
    # Add a new worksheet with name "Passif"
    passif = workbook.create_sheet("Passif")
    add_image(passif, image_path)
    # Get years from the BG affectée sheet
    bg_affectee = workbook["BG affectée"]
    year1 = bg_affectee["S4"].value
    year2 = bg_affectee["R4"].value  
    
    # Set values in specific cells
    cell_values = {
        "A3": "YAZAKI AUTOMOTIVE PRODUCTS TUNISIA Sarl",
        "B6": f"BILAN ARRETE au 31 mars {year1}",
        "B7": "(exprimé en Dinars Tunisiens)",
        "B9": "CAPITAUX PROPRES ET PASSIFS",
        "C9": "Notes",
        "B10": "Capitaux propres",
        "B12": "Capital libéré",
        "B13": "Reserve Legal",
        "B14": "Résultat reporté",
        "B16": "Total des capitaux propres avant résultat de l'exercice",
        "B18": "Résultat de l'exercice",
        "B20": "Total des capitaux propres avant affectation",
        "B22": "Emprunts à plus d'un an",
        "B23": "Provision pour risques et charges",
        "B25": "Provision pour risques et charges",
        "B27": "Fournisseurs et comptes rattachés",
        "B28": "Autres passifs courants",
        "B29": "Concours bancaires et autres passifs financiers",
        "B31": "Total des passifs courants",
        "B33": "TOTAL DES PASSIFS",
        "B35": "TOTAL DES CAPITAUX PROPRES ET PASSIFS",
        "C20": "III.7",
        "C22": "III.8",
        "C23": "III.9",
        "C27": "III.10",
        "C28": "III.11",
        "C29": "III.12",
        "D9": f"Solde au 31 mars {year1}",
        "E9": f"Solde au 31 mars {year2}"
    }
    
    for cell, value in cell_values.items():
        passif[cell] = value
    
    rows_for_long_formula = [12, 13, 14, 18, 22, 23, 27, 28, 29]
    rows_for_short_formula = [18]
    
    for row in rows_for_long_formula:
        if passif[f"B{row}"].value:
            passif[f"D{row}"] = f"=-SUMIF('BG affectée'!U:U,'Passif'!B{row},'BG affectée'!S:S)"
            passif[f"E{row}"] = f"=-SUMIF('BG affectée'!U:U,'Passif'!B{row},'BG affectée'!R:R)"
            passif[f"F{row}"] = f"=-SUMIF('BG affectée'!U:U,'Passif'!B{row},'BG affectée'!Q:Q)"
    
    for row in rows_for_short_formula:
        if passif[f"B{row}"].value:
            passif[f"D{row}"] = f"='P&L'!D37"
            passif[f"E{row}"] = f"='P&L'!E37"
    
    # Hide column F
    passif.column_dimensions['F'].hidden = True

    # Apply the necessary summary formulas
    summary_formulas = {
        "D16": "=SUM(D12:D14)",
        "E16": "=SUM(E12:E14)",
        "D20": "=SUM(D16:D18)",
        "E20": "=SUM(E16:E18)",
        "D25": "=SUM(D22:D23)",
        "E25": "=SUM(E22:E23)",
        "D31": "=SUM(D27:D29)",
        "E31": "=SUM(E27:E29)",
        "D33": "=D31+D25",
        "E33": "=E31+E25",
        "D35": "=D33+D20",
        "E35": "=E33+E20"
    }
    
    for cell, formula in summary_formulas.items():
        passif[cell] = formula
        passif[cell].font = Font(name="Times New Roman", bold=True)

    # Apply font properties to all cells
    font_times_new_roman = Font(name="Times New Roman", italic=False)
    for row in passif.iter_rows():
        for cell in row:
            cell.font = font_times_new_roman

    # Bold specific cells
    bold_cells = [
        "B14", "B22", "B20", "B6", "B9", "B10", "B16", "B25",
        "B31", "B33", "B35", "C9", "C20", "C21", "C22", "C27",
        "C28", "C29", "D33", "E33", "D35", "E35", "D31", "E31", "D26", "E26",
        "D25", "E25", "D20", "E20", "D16", "E16", "D9", "E9"
    ]
    for cell in bold_cells:
        passif[cell].font = Font(name="Times New Roman", bold=True)

    # Set column widths
    passif.column_dimensions['B'].width = 40
    passif.column_dimensions['C'].width = 10
    passif.column_dimensions['D'].width = 20
    passif.column_dimensions['E'].width = 20

    # Set alignment properties
    center_alignment = Alignment(horizontal="center", vertical="center")
    for cell in ["D9", "E9", "B6", "B7"]:
        passif[cell].alignment = center_alignment

    passif.merge_cells("B6:E6")
    passif.merge_cells("B7:E7")

    # Wrap text in headers
    for cell in ["D9", "E9", "B35"]:
        passif[cell].alignment = Alignment(wrap_text=True, horizontal="center")

    # Set gridlines visibility
    passif.sheet_view.showGridLines = False

    # Set number format for columns D and E to accounting format without symbol
    for col in ['D', 'E']:
        for row in range(1, passif.max_row + 1):
            passif[f"{col}{row}"].number_format = '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'

    return workbook

def P_L(workbook):
    image_path = r'1ksp_yazaki.png'
    # Add a new worksheet with name "P&L"
    pnl = workbook.create_sheet("P&L")
    add_image(pnl, image_path)
    # Get years from the BG affectée sheet
    bg_affectee = workbook["BG affectée"]
    year1 = bg_affectee["S4"].value
    year2 = bg_affectee["R4"].value 
    year3 = bg_affectee["Q4"].value 
    
    pnl["A3"] = "YAZAKI AUTOMOTIVE PRODUCTS TUNISIA Sarl"
    pnl["B6"] = f"ETAT DE RESULTAT ARRETE au 31 mars {year1}"
    pnl["B7"] = "(exprimé en Dinars Tunisiens)"
    pnl["B8"] = "Résultat"
    pnl["B10"] = "PRODUITS D'EXPLOITATION"
    pnl["B12"] = "Revenus"
    pnl["C12"] = "IV.1"
    pnl["B13"] = "Autres produits d'exploitation"
    pnl["C13"] = "IV.2"
    pnl["B15"] = "Total des produits d'exploitation"
    pnl["B17"] = "CHARGES D'EXPLOITATION"
    pnl["B19"] = "Variation de stocks de PF et encours"
    pnl["C19"] = "IV.3"
    pnl["B20"] = "Achats d'approvisionnements consommés"
    pnl["C20"] = "IV.3"
    pnl["B21"] = "Charges de personnel"
    pnl["C21"] = "IV.4"
    pnl["B22"] = "Dotations aux amortissements et aux provisions "
    pnl["C22"] = "IV.5"
    pnl["B23"] = "Autres Charges d'exploitation"
    pnl["C23"] = "IV.6"
    pnl["B25"] = "Total des charges d'exploitation"
    pnl["B27"] = "RESULTAT D'EXPLOITATION"
    pnl["B29"] = "Charges financières nettes"
    pnl["C29"] = "IV.7"
    pnl["B30"] = "Autres gains ordinaires"
    pnl["C30"] = "IV.8"
    pnl["B31"] = "Autres pertes ordinaires"
    pnl["C31"] = "IV.9"
    pnl["B33"] = "RESULTAT DES ACTIVITES ORDINAIRES AVANT IMPOT"
    pnl["B35"] = "Impôts sur les bénéfices"
    pnl["C35"] = "IV.10"
    pnl["B37"] = "RESULTAT NET DE L'EXERCICE"
    
    # Set headers for the years
    pnl["D8"] = f"Du 1er Avril {year2} au 31 mars {year1}"
    pnl["E8"] = f"Du 1er Avril {year3} au 31 mars {year2}"
    
    # Apply font properties to all cells
    font_times_new_roman = Font(name="Times New Roman", italic=False)
    for row in pnl.iter_rows():
        for cell in row:
            cell.font = font_times_new_roman
    
    # Bold specific cells
    bold_cells = [
        "A3", "B6", "B8", "B10", "B15", "B17", "B25", "B27", "B33", "B37",
        "D8", "E8", "B12", "B13", "B19", "B20", "B21", "B22", "B23", "B29", "B30", "B31", "B35",
        "C11","C12","C18","C19","C20","C21","C22","C23","C29","C30","C31","C35"
    ]
    for cell in bold_cells:
        pnl[cell].font = Font(name="Times New Roman", bold=True)
    
    # Set column widths
    pnl.column_dimensions['B'].width = 40
    pnl.column_dimensions['C'].width = 10
    pnl.column_dimensions['D'].width = 20
    pnl.column_dimensions['E'].width = 20
    
    # Set formulas in specific cells only if the corresponding cell in column B is not empty
    rows_for_negative_formula = [12, 13, 29, 30, 31, 35]
    rows_for_positive_formula = [19, 20, 21, 22, 23]
    
    for row in rows_for_negative_formula:
        if pnl[f"B{row}"].value:
            pnl[f"D{row}"] = f"=-SUMIF('BG affectée'!U:U,'P&L'!B{row},'BG affectée'!S:S)"
            pnl[f"E{row}"] = f"=-SUMIF('BG affectée'!U:U,'P&L'!B{row},'BG affectée'!R:R)"
            pnl[f"F{row}"] = f"=-SUMIF('BG affectée'!U:U,'P&L'!B{row},'BG affectée'!Q:Q)"
    
    for row in rows_for_positive_formula:
        if pnl[f"B{row}"].value:
            pnl[f"D{row}"] = f"=SUMIF('BG affectée'!U:U,'P&L'!B{row},'BG affectée'!S:S)"
            pnl[f"E{row}"] = f"=SUMIF('BG affectée'!U:U,'P&L'!B{row},'BG affectée'!R:R)"
            pnl[f"F{row}"] = f"=SUMIF('BG affectée'!U:U,'P&L'!B{row},'BG affectée'!Q:Q)"
    
    # Hide column F
    pnl.column_dimensions['F'].hidden = True
    # Apply the necessary summary formulas
    summary_formulas = {
        "D15": "=SUM(D12:D13)",
        "E15": "=SUM(E12:E13)",
        "F15": "=SUM(E12:E13)",
        "D25": "=SUM(D19:D23)",
        "E25": "=SUM(E19:E23)",
        "F25": "=SUM(E19:E23)",
        "D27": "=D15-D25",
        "E27": "=E15-E25",
        "F27": "=E15-E25",
        "D33": "=SUM(D27:D31)",
        "E33": "=SUM(E27:E31)",
        "F33": "=SUM(E27:E31)",
        "D37": "=D33+D35",
        "E37": "=E33+E35",
        "F37": "=E33+E35"
    }
    
    for cell, formula in summary_formulas.items():
        pnl[cell] = formula
        pnl[cell].font = Font(name="Times New Roman", bold=True)
            
    # Set alignment properties
    center_alignment = Alignment(horizontal="center", vertical="center")
    pnl["D9"].alignment = center_alignment
    pnl["E9"].alignment = center_alignment
    pnl["B6"].alignment = center_alignment
    pnl["B7"].alignment = center_alignment
    pnl.merge_cells("B6:E6")
    pnl.merge_cells("B7:E7")

    # Hide all columns first
    for col in range(1, pnl.max_column + 1):
        pnl.column_dimensions[openpyxl.utils.get_column_letter(col)].hidden = True

    # Unhide specific columns
    columns_to_keep = ['A','B', 'C', 'D', 'E']
    for col in columns_to_keep:
        pnl.column_dimensions[col].hidden = False

    # Hide all rows first
    for row in range(1, pnl.max_row + 1):
        pnl.row_dimensions[row].hidden = True

    # Unhide specific rows
    rows_to_keep = range(1, 38)  # Modify this range based on your specific requirements
    for row in rows_to_keep:
        pnl.row_dimensions[row].hidden = False

    # Wrap text
    pnl["D8"].alignment = Alignment(wrap_text=True, horizontal="center")
    pnl["E8"].alignment = Alignment(wrap_text=True, horizontal="center")
    pnl["B33"].alignment = Alignment(wrap_text=True, horizontal="center")

    # Set gridlines visibility
    pnl.sheet_view.showGridLines = False
    
    # Set number format for columns D and E to accounting format without symbol
    for col in ['D', 'E']:
        for row in range(1, pnl.max_row + 1):
            pnl[f"{col}{row}"].number_format = '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'

    return workbook
    
def bilan_passif_ang(workbook):
    image_path = r'1ksp_yazaki.png'
    # Add a new worksheet with name "Passif"
    passif = workbook.create_sheet("Passif ANG")
    add_image(passif, image_path)
    # Get years from the BG affectée sheet
    bg_affectee = workbook["BG affectée"]
    year1 = bg_affectee["S4"].value
    year2 = bg_affectee["R4"].value  
    
    # Set values in specific cells
    cell_values = {
        "A3": "YAZAKI AUTOMOTIVE PRODUCTS TUNISIA Sarl",
        "B6": f"BALANCE SHEET AS OF  March 31st,{year1}",
        "B7": "(in Tunisian Dinars)",
        "B9": "EQUITY AND LIABILITIES",
        "C9": "Notes",
        "B10": "Equity",
        "B12": "Share Capital",
        "B13": "Legal Reserve",
        "B14": "Retained earnings",
        "B16": "Total Equity before Result for the period",
        "B18": "Result for the period",
        "B20": "Total Equity ",
        "B22": "Long Term Borrowings",
        "B23": "Provisions for risks and charges ",
        "B25": "Total Non-Current Liabilities",
        "B27": "Accounts payable and related accounts",
        "B28": "Other current liabilities",
        "B29": "Bank borrowings and other liabilities",
        "B31": "Total Current Liabilities",
        "B33": "Total Liabilities",
        "B35": "TOTAL EQUITY AND LIABILITIES",
        "C20": "III.7",
        "C22": "III.8",
        "C23": "III.9",
        "C27": "III.10",
        "C28": "III.11",
        "C29": "III.12",
        "D9": f"Balance as of  March 31st,{year1}",
        "E9": f"Balance as of  March 31st,{year2}"
    }
    
    for cell, value in cell_values.items():
        passif[cell] = value
        
    for row in range(12, 45):
        if passif[f"B{row}"].value:
            passif[f"D{row}"] = f"=+'Passif'!D{row}"
            passif[f"E{row}"] = f"=+'Passif'!E{row}"    
    


    # Apply the necessary summary formulas
    summary_formulas = {
        "D16": "=SUM(D12:D14)",
        "E16": "=SUM(E12:E14)",
        "D20": "=SUM(D16:D18)",
        "E20": "=SUM(E16:E18)",
        "D25": "=SUM(D22:D23)",
        "E25": "=SUM(E22:E23)",
        "D31": "=SUM(D27:D29)",
        "E31": "=SUM(E27:E29)",
        "D33": "=D31+D25",
        "E33": "=E31+E25",
        "D35": "=D33+D20",
        "E35": "=E33+E20"
    }
    
    for cell, formula in summary_formulas.items():
        passif[cell] = formula
        passif[cell].font = Font(name="Times New Roman", bold=True)

    # Apply font properties to all cells
    font_times_new_roman = Font(name="Times New Roman", italic=False)
    for row in passif.iter_rows():
        for cell in row:
            cell.font = font_times_new_roman

    # Bold specific cells
    bold_cells = [
        "B14", "B22", "B20", "B6", "B9", "B10", "B16", "B25",
        "B31", "B33", "B35", "C9", "C20", "C21", "C22", "C27",
        "C28", "C29", "D33", "E33", "D35", "E35", "D31", "E31", "D26", "E26",
        "D25", "E25", "D20", "E20", "D16", "E16", "D9", "E9"
    ]
    for cell in bold_cells:
        passif[cell].font = Font(name="Times New Roman", bold=True)

    # Set column widths
    passif.column_dimensions['B'].width = 40
    passif.column_dimensions['C'].width = 10
    passif.column_dimensions['D'].width = 20
    passif.column_dimensions['E'].width = 20

    # Set alignment properties
    center_alignment = Alignment(horizontal="center", vertical="center")
    for cell in ["D9", "E9", "B6", "B7"]:
        passif[cell].alignment = center_alignment

    passif.merge_cells("B6:E6")
    passif.merge_cells("B7:E7")

    # Wrap text in headers
    for cell in ["D9", "E9", "B35"]:
        passif[cell].alignment = Alignment(wrap_text=True, horizontal="center")

    # Set gridlines visibility
    passif.sheet_view.showGridLines = False

    # Set number format for columns D and E to accounting format without symbol
    for col in ['D', 'E']:
        for row in range(1, passif.max_row + 1):
            passif[f"{col}{row}"].number_format = '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'

    return workbook

def P_L_ang(workbook):
    image_path = r'1ksp_yazaki.png'
    # Add a new worksheet with name "P&L"
    pnl = workbook.create_sheet("PL ANG")
    add_image(pnl, image_path)
    # Get years from the BG affectée sheet
    bg_affectee = workbook["BG affectée"]
    year1 = bg_affectee["S4"].value
    year2 = bg_affectee["R4"].value 
    year3 = bg_affectee["Q4"].value 
    
    pnl["A3"] = "YAZAKI AUTOMOTIVE PRODUCTS TUNISIA Sarl"
    pnl["B6"] = f"INCOME STATEMENT AS OF March 31st,{year1}"
    pnl["B7"] = "(in Tunisian Dinars)"
    pnl["B8"] = "Result"
    pnl["B10"] = "OPERATING REVENUES"
    pnl["B12"] = "Revenus"
    pnl["C12"] = "IV.1"
    pnl["B13"] = "Other Operating revenues"
    pnl["C13"] = "IV.2"
    pnl["B15"] = "TOTAL OPERATING REVENUES"
    pnl["B17"] = "OPERATING EXPENSES"
    pnl["B19"] = "Changes in inventories of finished goods and work in progress"
    pnl["C19"] = "IV.3"
    pnl["B20"] = "Consumed Purchases"
    pnl["C20"] = "IV.3"
    pnl["B21"] = "Payroll Expenses"
    pnl["C21"] = "IV.4"
    pnl["B22"] = "Provisions & Depreciation"
    pnl["C22"] = "IV.5"
    pnl["B23"] = "Other operating expenses"
    pnl["C23"] = "IV.6"
    pnl["B25"] = "Total Operating expenses"
    pnl["B27"] = "OPERATING RESULT"
    pnl["B29"] = "Net Financial Charges"
    pnl["C29"] = "IV.7"
    pnl["B30"] = "Other gains"
    pnl["C30"] = "IV.8"
    pnl["B31"] = "Other losses"
    pnl["C31"] = "IV.9"
    pnl["B33"] = "OPERATING RESULT BEFORE TAX"
    pnl["B35"] = "Income Tax"
    pnl["C35"] = "IV.10"
    pnl["B37"] = "NET RESULT FOR THE PERIOD"
    
    # Set headers for the years
    pnl["D8"] = f"From April 1st, {year2} to March 31st, {year1}"
    pnl["E8"] = f"From April 1st, {year3} to March 31st, {year2}"
    
    # Apply font properties to all cells
    font_times_new_roman = Font(name="Times New Roman", italic=False)
    for row in pnl.iter_rows():
        for cell in row:
            cell.font = font_times_new_roman
    
    # Bold specific cells
    bold_cells = [
        "A3", "B6", "B8", "B10", "B15", "B17", "B25", "B27", "B33", "B37",
        "D8", "E8", "B12", "B13", "B19", "B20", "B21", "B22", "B23", "B29", "B30", "B31", "B35",
        "C11","C12","C18","C19","C20","C21","C22","C23","C29","C30","C31","C35"
    ]
    for cell in bold_cells:
        pnl[cell].font = Font(name="Times New Roman", bold=True)
    
    # Set column widths
    pnl.column_dimensions['B'].width = 40
    pnl.column_dimensions['C'].width = 10
    pnl.column_dimensions['D'].width = 20
    pnl.column_dimensions['E'].width = 20
    
    # Set formulas in specific cells only if the corresponding cell in column B is not empty
    for row in range(12, 45):
        if pnl[f"B{row}"].value:
            pnl[f"D{row}"] = f"=+'P&L'!D{row}"
            pnl[f"E{row}"] = f"=+'P&L'!E{row}"
    # Apply the necessary summary formulas
    summary_formulas = {
        "D15": "=SUM(D12:D13)",
        "E15": "=SUM(E12:E13)",
        "D25": "=SUM(D19:D23)",
        "E25": "=SUM(E19:E23)",
        "D27": "=D15-D25",
        "E27": "=E15-E25",
        "D33": "=SUM(D27:D31)",
        "E33": "=SUM(E27:E31)",
        "D37": "=D33+D35",
        "E37": "=E33+E35"
    }
    
    for cell, formula in summary_formulas.items():
        pnl[cell] = formula
        pnl[cell].font = Font(name="Times New Roman", bold=True)
            
    # Set alignment properties
    center_alignment = Alignment(horizontal="center", vertical="center")
    pnl["D9"].alignment = center_alignment
    pnl["E9"].alignment = center_alignment
    pnl["B6"].alignment = center_alignment
    pnl["B7"].alignment = center_alignment
    pnl.merge_cells("B6:E6")
    pnl.merge_cells("B7:E7")

    # Hide all columns first
    for col in range(1, pnl.max_column + 1):
        pnl.column_dimensions[openpyxl.utils.get_column_letter(col)].hidden = True

    # Unhide specific columns
    columns_to_keep = ['A','B', 'C', 'D', 'E']
    for col in columns_to_keep:
        pnl.column_dimensions[col].hidden = False

    # Hide all rows first
    for row in range(1, pnl.max_row + 1):
        pnl.row_dimensions[row].hidden = True

    # Unhide specific rows
    rows_to_keep = range(1, 38)  # Modify this range based on your specific requirements
    for row in rows_to_keep:
        pnl.row_dimensions[row].hidden = False

    # Wrap text
    pnl["D8"].alignment = Alignment(wrap_text=True, horizontal="center")
    pnl["E8"].alignment = Alignment(wrap_text=True, horizontal="center")
    pnl["B33"].alignment = Alignment(wrap_text=True, horizontal="center")

    # Set gridlines visibility
    pnl.sheet_view.showGridLines = False
    
    # Set number format for columns D and E to accounting format without symbol
    for col in ['D', 'E']:
        for row in range(1, pnl.max_row + 1):
            pnl[f"{col}{row}"].number_format = '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'

    return workbook



def Prep_état(workbook):
    prep_etat_Fx = workbook.create_sheet("Prep état Fx")
    
    # Set headers
    headers = ["ACTIFS", "Solde au 31 mars 2024", "Solde au 31 mars 2023", "Solde au 31 mars 2023", "Diff N", "Diff N-1"]
    for col, header in enumerate(headers, start=2):
        prep_etat_Fx.cell(row=2, column=col, value=header)
    
    # Set column widths
    prep_etat_Fx.column_dimensions['B'].width = 35
    prep_etat_Fx.column_dimensions['C'].width = 20
    prep_etat_Fx.column_dimensions['D'].width = 20
    prep_etat_Fx.column_dimensions['E'].width = 20
    prep_etat_Fx.column_dimensions['F'].width = 20
    prep_etat_Fx.column_dimensions['G'].width = 20
    
    # Set range B3:B23 with specific values
    values = [
        "Immobilisations incorporelles", "Moins: Amortissements II", "Immobilisations corporelles", "Moins: Amortissements IC",
        "Immobilisations financières", "Moins : provisions IF", "Stocks", "Moins: Provisions pour dépréciation Stock",
        "Clients et comptes rattachés", "Moins : Provisions sur clients", "Autres actifs courants", 
        "Moins: Provisions pour dépréciation Actifs", "Liquidités et équivalents de liquidités", "Capital libéré", 
        "Résultat reporté", "Résultat de l'exercice", "Emprunts à plus d'un an", "Provision pour risques et charges",
        "Fournisseurs et comptes rattachés", "Autres passifs courants", "Concours bancaires et autres passifs financiers"
    ]
    for row, value in enumerate(values, start=3):
        prep_etat_Fx.cell(row=row, column=2, value=value)
    
    # Set range C3:C14 with formulas referencing Actif sheet
    actif_rows = [12, 13, 16, 17, 20, 21, 28, 29, 32, 33, 36, 37, 40]
    for i, row in enumerate(range(3, 16)):
        prep_etat_Fx[f"C{row}"] = f"='Actif'!D{actif_rows[i]}"
        prep_etat_Fx[f"D{row}"] = f"='Actif'!E{actif_rows[i]}"
        prep_etat_Fx[f"E{row}"] = f"=+SUMIF('BG affectée'!U:U,'Prep état Fx'!B{row},'BG affectée'!Q:Q)"
        
    passif_rows = [12, 14, 18, 22, 23, 27, 28, 29]
    for i, row in enumerate(range(16, 24)):
        if row == 18:
            prep_etat_Fx[f"C{row}"] = "='P&L'!D37"
            prep_etat_Fx[f"D{row}"] = "='P&L'!E37"
            prep_etat_Fx[f"E{row}"] = "='P&L'!F37"
        else:
            prep_etat_Fx[f"C{row}"] = f"='Passif'!D{passif_rows[i]}"
            prep_etat_Fx[f"D{row}"] = f"='Passif'!E{passif_rows[i]}"
            prep_etat_Fx[f"E{row}"] = f"=-SUMIF('BG affectée'!U:U,'Prep état Fx'!B{row},'BG affectée'!Q:Q)"
    
    # Set range F3:F23 and G3:G23 with specific formulas
    for row in range(3, 16):
        prep_etat_Fx[f"F{row}"] = f"=D{row}-C{row}"
        prep_etat_Fx[f"G{row}"] = f"=E{row}-D{row}"
    
    for row in range(15, 24):
        prep_etat_Fx[f"F{row}"] = f"=C{row}-D{row}"
        prep_etat_Fx[f"G{row}"] = f"=D{row}-E{row}"
        
    # Set number format for columns C to G to accounting format without decimals
    for col in ['C', 'D', 'E', 'F', 'G']:
        for row in range(3, 24):
            prep_etat_Fx[f"{col}{row}"].number_format = '_(* #,##0_);_(* (#,##0);_(* "-"??_);_(@_)'
    
    # Adding the table
    table_name = "Table2"
    if table_name in [t.name for t in prep_etat_Fx._tables]:
        prep_etat_Fx._tables.remove(prep_etat_Fx._tables[table_name])
    
    tab = openpyxl.worksheet.table.Table(displayName=table_name, ref="B2:G23")
    style = openpyxl.worksheet.table.TableStyleInfo(name="TableStyleLight14", showFirstColumn=False,
                                                    showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tab.tableStyleInfo = style
    prep_etat_Fx.add_table(tab)
    
    # Additional rows and VLOOKUP formulas
    prep_etat_Fx["A41"] = 8493000
    prep_etat_Fx["A42"] = 8496000
    prep_etat_Fx["A43"] = 8497000
    prep_etat_Fx["A44"] = "X849400"
    prep_etat_Fx["A45"] = "X849700"
    prep_etat_Fx["A46"] = "X849900"
    prep_etat_Fx["A47"] = "X849900"
    
    prep_etat_Fx["B39"] = " +/- Value de cession"
    
    prep_etat_Fx["B41"] = "FA Gains Proceeds"
    prep_etat_Fx["B42"] = "FA Losses Proceeds"
    prep_etat_Fx["B43"] = "Fixed Asset Losses - Net Book Value"
    prep_etat_Fx["B44"] = "FA Gains NBV"
    prep_etat_Fx["B45"] = "FA Losses NBV"
    prep_etat_Fx["B46"] = "FA Disposal Clearing"
    prep_etat_Fx["B47"] = "Fixed Asset Gains - Net Book Value"
    prep_etat_Fx["C40"] = "M"
    prep_etat_Fx["D40"] = "M-1"
    
    prep_etat_Fx["C48"] = "=SUM(C41:C47)"
    prep_etat_Fx["D48"] = "=SUM(D41:D47)"
    
    additional_rows = [41, 42, 43, 44, 45, 46, 47]
    for row in additional_rows:
        if prep_etat_Fx[f"A{row}"].value:
            prep_etat_Fx[f"C{row}"] = f"=+VLOOKUP(A{row},'BG affectée'!D:S,16,0)"
            prep_etat_Fx[f"D{row}"] = f"=+VLOOKUP(A{row},'BG affectée'!D:S,15,0)"
    
    # Set gridlines visibility
    prep_etat_Fx.sheet_view.showGridLines = False
    
    return workbook



        

def get_download_link(file_path, filename: str):
    with open(file_path, 'rb') as f:
        b64 = base64.b64encode(f.read()).decode()
    return f'<a href="data:application/octet-stream;base64,{b64}" download="{filename}">Download {filename}</a>'

def sheet_exists(file_path, sheet_name):
    try:
        workbook = openpyxl.load_workbook(file_path, read_only=True)
        return sheet_name in workbook.sheetnames
    except Exception as e:
        st.error(f"Failed to check if sheet exists: {e}")
        return False

def add_sheet_to_excel(file_path, df, sheet_name):
    try:
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    except Exception as e:
        st.error(f"Failed to add sheet to Excel file: {e}")

def display_welcome_page() -> None:
    """Display the welcome page content."""
    st.title("Welcome to the Yazaki Tunisia Finance Assistant")
    st.image(IMAGE_PATH, use_column_width=True)
    st.markdown("""
        ### Overview
         Welcome to the **Yazaki Finance Assistant**, your comprehensive solution for managing and analyzing financial data with precision and ease. Whether you're a seasoned finance professional or an emerging analyst, our platform is designed to streamline your financial operations and provide you with deep insights.

        ###  Features at a Glance
        - **Upload & Process Excel Files**: Easily upload your financial data files and let our advanced tools handle the processing.
        - **Interactive Dashboards**: Visualize your data through intuitive dashboards that reveal key trends and insights.
        - **AI Financial Assistant**: Utilize our AI-powered assistant to get expert answers and recommendations for your financial queries.
        - **Trend Analysis**: Examine financial trends over time to support strategic decision-making.

        ###  Getting Started
        Navigate through the sidebar to access different sections of our app:
        - **Upload & Process Excel**: Begin by uploading your financial data for seamless processing.
        - **KPIS**: View and analyze key performance indicators and essential financial metrics.
        - **AI Assistant**: Interact with our AI assistant for tailored financial advice and insights.
        - **Trend Analysis**: Explore financial trends over various periods to make informed decisions.

        ###  About Yazaki
        Yazaki is a global leader in automotive parts and systems, committed to innovation and excellence. Our solutions drive the future of the automotive industry, ensuring top quality and reliability.

        Experience the power of the Yazaki Finance Assistant – transforming data into actionable insights, and helping you achieve financial clarity and success. 🎉
    """)
    st.video(WELCOME_VIDEO_URL)

def display_upload_page(lottie_upload: dict) -> None:
    st.title("Upload & Process Excel")
    
    if lottie_upload:
        st_lottie(lottie_upload, height=300, key="upload")
    else:
        st.warning("Lottie animation could not be loaded.")

    uploaded_file = st.file_uploader("Upload an Excel file", type=["xlsx", "csv"])
    TARGET_FILE_PATH = r'rej.xlsx'
    if uploaded_file:
        st.write("### File Details")
        st.write("**Filename:**", uploaded_file.name)
        
        # Ensure df is assigned before being used
        if uploaded_file.type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet':
            df = pd.read_excel(uploaded_file)
        else:
            df = pd.read_csv(uploaded_file)
        
        st.write("### Preview of Uploaded File")
        st.dataframe(df.head())

        if st.button("Process and Download Excel File"):
            with st.spinner('Processing...'):
                sheet_already_exists = sheet_exists(TARGET_FILE_PATH, 'BG SAP')
                
                if not sheet_already_exists:
                    add_sheet_to_excel(TARGET_FILE_PATH, df, 'BG SAP')
                    st.success(f"Data added as a new sheet to {TARGET_FILE_PATH}")
                else:
                    st.warning(f"Sheet 'BG SAP' already exists in {TARGET_FILE_PATH}. Skipping addition.")
                
                workbook = process_excel(TARGET_FILE_PATH)
                if workbook:
                    # Run additional processing functions
                    workbook = bilan_actif(workbook)
                    workbook = bilan_actif_ang(workbook)
                    workbook = P_L(workbook)
                    workbook = P_L_ang(workbook)
                    workbook = bilan_passif(workbook)
                    workbook = bilan_passif_ang(workbook)
                    workbook = Prep_état(workbook)
                    
                    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
                        workbook.save(tmp.name)
                        tmp_path = tmp.name
                    # Read the processed file back into a DataFrame
                    processed_df = pd.read_excel(tmp_path)
                    st.success('File processed successfully!')
                    st.markdown(get_download_link(tmp_path, "BG_file.xlsx"), unsafe_allow_html=True)
                    st.session_state['processed_df'] = processed_df
                    
def load_and_clean_data(df):
    df.columns = df.columns.astype(str)
    if 'Affectation rapport ' not in df.columns:
        st.error("The column 'Affectation rapport ' is missing from the uploaded data.")
        return pd.DataFrame()
    solde_columns = [col for col in df.columns if 'Solde' in col]
    unpivoted_data = df.melt(id_vars=['Affectation rapport '], value_vars=solde_columns, var_name='Year', value_name='Balance')
    unpivoted_data['Year'] = unpivoted_data['Year'].str.extract(r'(\d{4})')
    unpivoted_data['Balance'] = pd.to_numeric(unpivoted_data['Balance'], errors='coerce')
    return unpivoted_data


def display_dashboards_page() -> None:
    """Display the dashboards page content."""
    
    # Load and process the Excel file
    file = r"rej12.xlsx"
    #df_clean = load_and_clean_data(df)
    # Load data and clean it
    df_clean = load_and_clean_data(pd.read_excel(file,header=3))

    # Classification lists
    Assets = [
        "Immobilisations incorporelles", "Moins: Amortissements II", 
        "Immobilisations corporelles", "Moins: Amortissements IC", 
        "Immobilisations financières", "Moins : provisions IF", 
        "Autres actifs non courants", "Stocks", "Moins: Provisions pour dépréciation Stock", 
        "Clients et comptes rattachés", "Moins : Provisions sur clients", 
        "Autres actifs courants", "Moins: Provisions pour dépréciation Actifs", 
        "Liquidités et équivalents de liquidités"
    ]
    equity_and_liabilities = [
        "Capital libéré", "Reserve Legal", "Résultat reporté", "Résultat de l'exercice", 
        "Emprunts à plus d'un an", "Provision pour risques et charges", 
        "Fournisseurs et comptes rattachés", "Autres passifs courants", 
        "Concours bancaires et autres passifs financiers"
    ]
    profit_and_loss = [
        "Revenus", "Autres produits d'exploitation", "Variation de stocks de PF et encours", 
        "Achats d'approvisionnements consommés", "Charges de personnel", 
        "Dotations aux amortissements et aux provisions ", "Autres Charges d'exploitation",  
        "Charges financières nettes", "Autres gains ordinaires", "Autres pertes ordinaires", 
        "Impôts sur les bénéfices"
    ]

    # Classification function
    def classify_accounts(value):
        if value in Assets:
            return "Assets"
        elif value in equity_and_liabilities:
            return "Equity and Liabilities"
        elif value in profit_and_loss:
            return "Profit and Loss"

    

    df_clean['Classification'] = df_clean['Affectation rapport '].apply(classify_accounts)

    # Sidebar for year selection
    st.sidebar.header("Select Year")
    years = sorted(df_clean['Year'].unique(), reverse=True)
    selected_year = st.sidebar.selectbox('Select Year', years)

    filtered_df = df_clean[df_clean['Year'] == selected_year]

    # Calculate financial metrics
    operating_revenues = (
        filtered_df[filtered_df['Affectation rapport '] == 'Revenus']['Balance'].sum() + 
        filtered_df[filtered_df['Affectation rapport '] == "Autres produits d'exploitation"]['Balance'].sum()
    ) * -1

    total_operating_expenses = (
        filtered_df[filtered_df['Affectation rapport '] == 'Variation de stocks de PF et encours']['Balance'].sum() +
        filtered_df[filtered_df['Affectation rapport '] == "Achats d'approvisionnements consommés"]['Balance'].sum() +
        filtered_df[filtered_df['Affectation rapport '] == 'Charges de personnel']['Balance'].sum() +
        filtered_df[filtered_df['Affectation rapport '] == 'Dotations aux amortissements et aux provisions ']['Balance'].sum() +
        filtered_df[filtered_df['Affectation rapport '] == "Autres Charges d'exploitation"]['Balance'].sum()
    )

    other_gains_and_losses = (
        filtered_df[filtered_df['Affectation rapport '] == 'Charges financières nettes']['Balance'].sum() +
        filtered_df[filtered_df['Affectation rapport '] == 'Autres gains ordinaires']['Balance'].sum() +
        filtered_df[filtered_df['Affectation rapport '] == 'Autres pertes ordinaires']['Balance'].sum() +
        filtered_df[filtered_df['Affectation rapport '] == 'Impôts sur les bénéfices']['Balance'].sum()
    ) * -1

    total_non_current_liabilities = (
        filtered_df[filtered_df['Affectation rapport '] == "Emprunts à plus d'un an"]['Balance'].sum() +
        filtered_df[filtered_df['Affectation rapport '] == 'Provision pour risques et charges']['Balance'].sum()
    ) * -1

    total_current_liabilities = (
        filtered_df[filtered_df['Affectation rapport '].str.lower() == 'autres passifs courants'.lower()]['Balance'].sum() +
        filtered_df[filtered_df['Affectation rapport '] == 'Fournisseurs et comptes rattachés']['Balance'].sum() +
        filtered_df[filtered_df['Affectation rapport '] == 'Concours bancaires et autres passifs financiers']['Balance'].sum()
    ) * -1

    total_liabilities = total_non_current_liabilities + total_current_liabilities

    total_equity = (
        -1 * filtered_df[filtered_df['Affectation rapport '] == 'Capital libéré']['Balance'].sum() +
        -1 * filtered_df[filtered_df['Affectation rapport '] == 'Reserve Legal']['Balance'].sum() +
        -1 * filtered_df[filtered_df['Affectation rapport '] == 'Résultat reporté']['Balance'].sum() +
        (operating_revenues - total_operating_expenses + other_gains_and_losses)
    )

    total_profit_and_loss = operating_revenues - total_operating_expenses + other_gains_and_losses

    total_equity_and_liabilities = total_equity + total_liabilities

    total_assets = filtered_df[filtered_df['Classification'] == 'Assets']['Balance'].sum()

    # Calculate financial ratios
    current_assets_total = (
        filtered_df[filtered_df['Affectation rapport '] == 'Stocks']['Balance'].sum() +
        filtered_df[filtered_df['Affectation rapport '] == 'Clients et comptes rattachés']['Balance'].sum() +
        filtered_df[filtered_df['Affectation rapport '] == 'Autres actifs courants']['Balance'].sum() +
        filtered_df[filtered_df['Affectation rapport '] == 'Liquidités et équivalents de liquidités']['Balance'].sum()
    )

    current_ratio = current_assets_total / total_current_liabilities if total_current_liabilities != 0 else None
    debt_to_equity_ratio = total_liabilities / total_equity if total_equity != 0 else None
    debt_ratio = total_liabilities / total_assets if total_assets != 0 else None
    equity_ratio = total_equity / total_assets if total_assets != 0 else None

    gross_profit = operating_revenues - total_operating_expenses
    net_income = gross_profit + other_gains_and_losses

    gross_profit_margin = (gross_profit / operating_revenues) if operating_revenues != 0 else None
    operating_profit_margin = (gross_profit / operating_revenues) if operating_revenues != 0 else None
    net_profit_margin = (net_income / operating_revenues) if operating_revenues != 0 else None

    return_on_assets = (net_income / total_assets) if total_assets != 0 else None
    return_on_equity = (net_income / total_equity) if total_equity != 0 else None

    inventory = filtered_df[filtered_df['Affectation rapport '] == 'Stocks']['Balance'].sum()
    inventory_turnover = (total_operating_expenses / inventory) if inventory != 0 else None

    accounts_receivable = filtered_df[filtered_df['Affectation rapport '] == 'Clients et comptes rattachés']['Balance'].sum()
    accounts_payable = filtered_df[filtered_df['Affectation rapport '] == 'Fournisseurs et comptes rattachés']['Balance'].sum()

    receivables_turnover = (operating_revenues / accounts_receivable) if accounts_receivable != 0 else None
    payables_turnover = (total_operating_expenses / accounts_payable) if accounts_payable != 0 else None

    # CSS for card styling
    st.markdown("""
        <style>
        .card {
            background: #f8f9fa;
            border-radius: 10px;
            box-shadow: 0 4px 8px 0 rgba(0,0,0,0.2);
            transition: 0.3s;
            padding: 20px;
            margin: 10px;
            text-align: center;
            width: 400px;
            height: 200px;
            display: flex;
            flex-direction: column;
            justify-content: center;
            align-items: center;
        }
        .card:hover {
            box-shadow: 0 8px 16px 0 rgba(0,0,0,0.2);
        }
        .container {
            display: flex;
            flex-wrap: wrap;
            justify-content: space-around;
        }
        .container > .card {
            flex: 1 1 calc(33% - 20px);
            margin: 10px;
        }
        .header-image {
            display: block;
            margin: 20px auto 0;
            width: 200px;
        }
        .interpretation-card {
            background: #f8f9fa;
            border-radius: 10px;
            box-shadow: 0 4px 8px 0 rgba(0,0,0,0.2);
            transition: 0.3s;
            padding: 20px;
            margin: 10px;
            text-align: left;
            white-space: pre-wrap;
            font-family: 'Arial', sans-serif;
            font-size: 16px;
            line-height: 1.6;
        }
        .interpretation-container {
            display: flex;
            justify-content: center;
        }
        .special-card {
            background: rgba(173, 216, 230, 0.5); /* Light blue transparent background */
        }    
        </style>
    """, unsafe_allow_html=True)

    # Function to create a card
    def create_card(title, value, description=None):
        return f"""
        <div class="card">
            <h3>{title}</h3>
            <h2>{value}</h2>
            {f'<p>{description}</p>' if description else ''}
        </div>
        """

    # Display header image
    image_path = r'1ksp_yazaki.png'
    image = Image.open(image_path)
    st.image(image, width=500)
    buffered = BytesIO()
    image.save(buffered, format="PNG")

    # Display Cards for Classification Totals
    def create_special_card(title, value):
     return f"""
    <div class="card special-card">
        <h3>{title}</h3>
        <h2>{value}</h2>
    </div>
    """
    #st.header(f'Totals for {selected_year} in TND')
    st.markdown(f"""
    <h2 style='color: #190478;font-size: 45px;'>Totals for The Fiscal year of {selected_year} in TND</h2>
    """, unsafe_allow_html=True)
    totals_cols = st.columns(3)
    totals_cols[0].markdown(create_special_card("Total Result", f"{total_profit_and_loss:,.2f}"), unsafe_allow_html=True)
    totals_cols[1].markdown(create_special_card("Total Equity", f"{total_equity:,.2f}"), unsafe_allow_html=True)
    totals_cols[2].markdown(create_special_card("Total Assets", f"{total_assets:,.2f}"), unsafe_allow_html=True)
    
    

    # Financial Ratios
    #st.header('Financial Ratios')
    st.markdown("""
    <h2 style='color: #190478;font-size: 45px;'>Financial Ratios</h2>
    """, unsafe_allow_html=True)
    ratios_cols = st.columns(4)
    ratios_cols[0].markdown(create_card("Current Ratio", f"{current_ratio:.2f}" if current_ratio else "N/A", "Formula: Current Assets / Current Liabilities"), unsafe_allow_html=True)
    ratios_cols[1].markdown(create_card("Debt to Equity Ratio", f"{debt_to_equity_ratio:.2f}" if debt_to_equity_ratio else "N/A", "Formula: Total Liabilities / Total Equity"), unsafe_allow_html=True)
    ratios_cols[2].markdown(create_card("Gross Profit Margin", f"{gross_profit_margin:.2%}" if gross_profit_margin else "N/A", "Formula: Gross Profit / Revenue"), unsafe_allow_html=True)
    ratios_cols[3].markdown(create_card("Operating Profit Margin", f"{operating_profit_margin:.2%}" if operating_profit_margin else "N/A", "Formula: Operating Profit / Revenue"), unsafe_allow_html=True)
    ratios_cols[0].markdown(create_card("Net Profit Margin", f"{net_profit_margin:.2%}" if net_profit_margin else "N/A", "Formula: Net Income / Revenue"), unsafe_allow_html=True)
    ratios_cols[1].markdown(create_card("Return on Assets (ROA)", f"{return_on_assets:.2%}" if return_on_assets else "N/A", "Formula: Net Income / Total Assets"), unsafe_allow_html=True)
    ratios_cols[2].markdown(create_card("Return on Equity (ROE)", f"{return_on_equity:.2%}" if return_on_equity else "N/A", "Formula: Net Income / Total Equity"), unsafe_allow_html=True)
    ratios_cols[3].markdown(create_card("Inventory Turnover", f"{inventory_turnover:.2f}" if inventory_turnover else "N/A", "Formula: Cost of Goods Sold / Average Inventory"), unsafe_allow_html=True)
    ratios_cols[0].markdown(create_card("Receivables Turnover", f"{receivables_turnover:.2f}" if receivables_turnover else "N/A", "Formula: Net Credit Sales / Average Accounts Receivable"), unsafe_allow_html=True)
    ratios_cols[1].markdown(create_card("Payables Turnover", f"{payables_turnover:.2f}" if payables_turnover else "N/A", "Formula: Total Supplier Purchases / Average Accounts Payable"), unsafe_allow_html=True)
    ratios_cols[2].markdown(create_card("Debt Ratio", f"{debt_ratio:.2%}" if debt_ratio else "N/A", "Formula: Total Liabilities / Total Assets"), unsafe_allow_html=True)
    ratios_cols[3].markdown(create_card("Equity Ratio", f"{equity_ratio:.2%}" if equity_ratio else "N/A", "Formula: Total Equity / Total Assets"), unsafe_allow_html=True)

    # LangChain setup
    llm = ChatOpenAI(api_key=openai.api_key, model="gpt-3.5-turbo")

    # Prepare the ratios for interpretation
    ratios_text = f"""
    Current Ratio: {current_ratio:.2f}
    Debt to Equity Ratio: {debt_to_equity_ratio:.2f}
    Gross Profit Margin: {gross_profit_margin:.2%}
    Operating Profit Margin: {operating_profit_margin:.2%}
    Net Profit Margin: {net_profit_margin:.2%}
    Return on Assets (ROA): {return_on_assets:.2%}
    Return on Equity (ROE): {return_on_equity:.2%}
    Inventory Turnover: {inventory_turnover:.2f}
    Receivables Turnover: {receivables_turnover:.2f}
    Payables Turnover: {payables_turnover:.2f}
    Debt Ratio: {debt_ratio:.2%}
    Equity Ratio: {equity_ratio:.2%}
    """

    interpretation_prompt = f"""
    As a financial analyst, interpret the following financial ratios for Yazaki and provide a deep analysis with suggestions for the company's financial health:

    {ratios_text}

    Your analysis should cover liquidity, profitability, leverage, and efficiency aspects, along with any potential risks and recommendations for improvement.
    """

    interpretation_response = llm.invoke(interpretation_prompt)

    st.header('Interpretation and Suggestions')

    interpretation_content = interpretation_response.content

    st.markdown(f"""
    <div class="interpretation-container">
        <div class="interpretation-card">
            {interpretation_content}
        </div>
    </div>
    """, unsafe_allow_html=True)
        

def calculate_financial_ratios(df):
    operating_revenues = (
        df[df['Affectation rapport '] == 'Revenus']['Balance'].sum() + 
        df[df['Affectation rapport '] == "Autres produits d'exploitation"]['Balance'].sum()
    ) * -1

    total_operating_expenses = (
        df[df['Affectation rapport '] == 'Variation de stocks de PF et encours']['Balance'].sum() +
        df[df['Affectation rapport '] == "Achats d'approvisionnements consommés"]['Balance'].sum() +
        df[df['Affectation rapport '] == 'Charges de personnel']['Balance'].sum() +
        df[df['Affectation rapport '] == 'Dotations aux amortissements et aux provisions ']['Balance'].sum() +
        df[df['Affectation rapport '] == "Autres Charges d'exploitation"]['Balance'].sum()
    )

    other_gains_and_losses = (
        df[df['Affectation rapport '] == 'Charges financières nettes']['Balance'].sum() +
        df[df['Affectation rapport '] == 'Autres gains ordinaires']['Balance'].sum() +
        df[df['Affectation rapport '] == 'Autres pertes ordinaires']['Balance'].sum() +
        df[df['Affectation rapport '] == 'Impôts sur les bénéfices']['Balance'].sum()
    ) * -1

    total_non_current_liabilities = (
        df[df['Affectation rapport '] == "Emprunts à plus d'un an"]['Balance'].sum() +
        df[df['Affectation rapport '] == 'Provision pour risques et charges']['Balance'].sum()
    ) * -1

    total_current_liabilities = (
        df[df['Affectation rapport '].str.lower() == 'autres passifs courants'.lower()]['Balance'].sum() +
        df[df['Affectation rapport '] == 'Fournisseurs et comptes rattachés']['Balance'].sum() +
        df[df['Affectation rapport '] == 'Concours bancaires et autres passifs financiers']['Balance'].sum()
    ) * -1

    total_liabilities = total_non_current_liabilities + total_current_liabilities

    total_equity = (
        -1 * df[df['Affectation rapport '] == 'Capital libéré']['Balance'].sum() +
        -1 * df[df['Affectation rapport '] == 'Reserve Legal']['Balance'].sum() +
        -1 * df[df['Affectation rapport '] == 'Résultat reporté']['Balance'].sum() +
        (operating_revenues - total_operating_expenses + other_gains_and_losses)
    )

    total_assets = df[df['Classification'] == 'Assets']['Balance'].sum()

    current_assets_total = (
        df[df['Affectation rapport '] == 'Stocks']['Balance'].sum() +
        df[df['Affectation rapport '] == 'Clients et comptes rattachés']['Balance'].sum() +
        df[df['Affectation rapport '] == 'Autres actifs courants']['Balance'].sum() +
        df[df['Affectation rapport '] == 'Liquidités et équivalents de liquidités']['Balance'].sum()
    )

    gross_profit = operating_revenues - total_operating_expenses
    net_income = gross_profit + other_gains_and_losses

    gross_profit_margin = (gross_profit / operating_revenues) if operating_revenues != 0 else None
    return_on_assets = (net_income / total_assets) if total_assets != 0 else None
    return_on_equity = (net_income / total_equity) if total_equity != 0 else None
    debt_ratio = total_liabilities / total_assets if total_assets != 0 else None
    equity_ratio = total_equity / total_assets if total_assets != 0 else None

    return pd.Series({
        'Gross Profit Margin': gross_profit_margin,
        'Return on Assets (ROA)': return_on_assets,
        'Return on Equity (ROE)': return_on_equity,
        'Debt Ratio': debt_ratio,
        'Equity Ratio': equity_ratio
    })

def classify_accounts(value):
    Assets = [
        "Immobilisations incorporelles", "Moins: Amortissements II", 
        "Immobilisations corporelles", "Moins: Amortissements IC", 
        "Immobilisations financières", "Moins : provisions IF", 
        "Autres actifs non courants", "Stocks", "Moins: Provisions pour dépréciation Stock", 
        "Clients et comptes rattachés", "Moins : Provisions sur clients", 
        "Autres actifs courants", "Moins: Provisions pour dépréciation Actifs", 
        "Liquidités et équivalents de liquidités"
    ]
    equity_and_liabilities = [
        "Capital libéré", "Reserve Legal", "Résultat reporté", "Résultat de l'exercice", 
        "Emprunts à plus d'un an", "Provision pour risques et charges", 
        "Fournisseurs et comptes rattachés", "Autres passifs courants", 
        "Concours bancaires et autres passifs financiers"
    ]
    profit_and_loss = [
        "Revenus", "Autres produits d'exploitation", "Variation de stocks de PF et encours", 
        "Achats d'approvisionnements consommés", "Charges de personnel", 
        "Dotations aux amortissements et aux provisions ", "Autres Charges d'exploitation",  
        "Charges financières nettes", "Autres gains ordinaires", "Autres pertes ordinaires", 
        "Impôts sur les bénéfices"
    ]
    
    if value in Assets:
        return "Assets"
    elif value in equity_and_liabilities:
        return "Equity and Liabilities"
    elif value in profit_and_loss:
        return "Profit and Loss"

def display_trend_analysis_page() -> None:
    file = r"rej12.xlsx"
    df_clean = load_and_clean_data(pd.read_excel(file, header=3))

    if df_clean.empty:
        return

    # Apply classification
    df_clean['Classification'] = df_clean['Affectation rapport '].apply(classify_accounts)

    # Debug check: Ensure 'Classification' column is created
    if 'Classification' not in df_clean.columns:
        st.error("Failed to create 'Classification' column.")
        return
    
    st.subheader('Trend Analysis of Financial Metrics Over Time')

    # Calculate metrics for each year using the same formulas
    metrics_by_year = df_clean.groupby('Year', group_keys=False).apply(lambda df: pd.Series({
        'Operating Revenues': (
            df[df['Affectation rapport '] == 'Revenus']['Balance'].sum() + 
            df[df['Affectation rapport '] == "Autres produits d'exploitation"]['Balance'].sum()
        ) * -1,
        'Total Operating Expenses': (
            df[df['Affectation rapport '] == 'Variation de stocks de PF et encours']['Balance'].sum() +
            df[df['Affectation rapport '] == "Achats d'approvisionnements consommés"]['Balance'].sum() +
            df[df['Affectation rapport '] == 'Charges de personnel']['Balance'].sum() +
            df[df['Affectation rapport '] == 'Dotations aux amortissements et aux provisions ']['Balance'].sum() +
            df[df['Affectation rapport '] == "Autres Charges d'exploitation"]['Balance'].sum()
        ),
        'Other Gains and Losses': (
            df[df['Affectation rapport '] == 'Charges financières nettes']['Balance'].sum() +
            df[df['Affectation rapport '] == 'Autres gains ordinaires']['Balance'].sum() +
            df[df['Affectation rapport '] == 'Autres pertes ordinaires']['Balance'].sum() +
            df[df['Affectation rapport '] == 'Impôts sur les bénéfices']['Balance'].sum()
        ) * -1,
        'Total Non-Current Liabilities': (
            df[df['Affectation rapport '] == "Emprunts à plus d'un an"]['Balance'].sum() +
            df[df['Affectation rapport '] == 'Provision pour risques et charges']['Balance'].sum()
        ) * -1,
        'Total Current Liabilities': (
            df[df['Affectation rapport '].str.lower() == 'autres passifs courants'.lower()]['Balance'].sum() +
            df[df['Affectation rapport '] == 'Fournisseurs et comptes rattachés']['Balance'].sum() +
            df[df['Affectation rapport '] == 'Concours bancaires et autres passifs financiers']['Balance'].sum()
        ) * -1,
        'Total Assets': df[df['Classification'] == 'Assets']['Balance'].sum(),
        'Total Equity': (
            -1 * df[df['Affectation rapport '] == 'Capital libéré']['Balance'].sum() +
            -1 * df[df['Affectation rapport '] == 'Reserve Legal']['Balance'].sum() +
            -1 * df[df['Affectation rapport '] == 'Résultat reporté']['Balance'].sum() +
            (
                (df[df['Affectation rapport '] == 'Revenus']['Balance'].sum() + 
                 df[df['Affectation rapport '] == "Autres produits d'exploitation"]['Balance'].sum()) * -1
                -
                (
                    df[df['Affectation rapport '] == 'Variation de stocks de PF et encours']['Balance'].sum() +
                    df[df['Affectation rapport '] == "Achats d'approvisionnements consommés"]['Balance'].sum() +
                    df[df['Affectation rapport '] == 'Charges de personnel']['Balance'].sum() +
                    df[df['Affectation rapport '] == 'Dotations aux amortissements et aux provisions ']['Balance'].sum() +
                    df[df['Affectation rapport '] == "Autres Charges d'exploitation"]['Balance'].sum()
                ) +
                (
                    df[df['Affectation rapport '] == 'Charges financières nettes']['Balance'].sum() +
                    df[df['Affectation rapport '] == 'Autres gains ordinaires']['Balance'].sum() +
                    df[df['Affectation rapport '] == 'Autres pertes ordinaires']['Balance'].sum() +
                    df[df['Affectation rapport '] == 'Impôts sur les bénéfices']['Balance'].sum()
                ) * -1
            )
        )
    })).reset_index()

    metrics_by_year['Total Liabilities'] = metrics_by_year['Total Non-Current Liabilities'] + metrics_by_year['Total Current Liabilities']
    metrics_by_year['Total Profit and Loss'] = metrics_by_year['Operating Revenues'] - metrics_by_year['Total Operating Expenses'] + metrics_by_year['Other Gains and Losses']
    metrics_by_year['Total Equity'] = metrics_by_year['Total Equity'] 
    content = ["Total Assets", "Total Equity", "Total Profit and Loss"]
    trend_data = metrics_by_year[['Year'] + content].melt(id_vars=['Year'], var_name='Metric', value_name='Value')

    # Create the trend chart
    trend_chart = alt.Chart(trend_data).mark_line(point=True).encode(
        x='Year:N',
        y='Value:Q',
        color='Metric:N',
        tooltip=['Year', 'Metric', 'Value']
    ).properties(
        width=400,
        height=300,
        title="Trend Analysis of Financial Metrics Over the Years"
    ).configure_axis(
        labelFontSize=12,
        titleFontSize=14
    ).configure_title(
        fontSize=18
    )
    st.altair_chart(trend_chart, use_container_width=True)

    st.subheader('Trend Analysis of Financial Ratios Over Time')

    # Calculate the financial ratios over time
    def calculate_financial_ratios(df):
        operating_revenues = (
            df[df['Affectation rapport '] == 'Revenus']['Balance'].sum() + 
            df[df['Affectation rapport '] == "Autres produits d'exploitation"]['Balance'].sum()
        ) * -1

        total_operating_expenses = (
            df[df['Affectation rapport '] == 'Variation de stocks de PF et encours']['Balance'].sum() +
            df[df['Affectation rapport '] == "Achats d'approvisionnements consommés"]['Balance'].sum() +
            df[df['Affectation rapport '] == 'Charges de personnel']['Balance'].sum() +
            df[df['Affectation rapport '] == 'Dotations aux amortissements et aux provisions ']['Balance'].sum() +
            df[df['Affectation rapport '] == "Autres Charges d'exploitation"]['Balance'].sum()
        )

        other_gains_and_losses = (
            df[df['Affectation rapport '] == 'Charges financières nettes']['Balance'].sum() +
            df[df['Affectation rapport '] == 'Autres gains ordinaires']['Balance'].sum() +
            df[df['Affectation rapport '] == 'Autres pertes ordinaires']['Balance'].sum() +
            df[df['Affectation rapport '] == 'Impôts sur les bénéfices']['Balance'].sum()
        ) * -1

        total_non_current_liabilities = (
            df[df['Affectation rapport '] == "Emprunts à plus d'un an"]['Balance'].sum() +
            df[df['Affectation rapport '] == 'Provision pour risques et charges']['Balance'].sum()
        ) * -1

        total_current_liabilities = (
            df[df['Affectation rapport '].str.lower() == 'autres passifs courants'.lower()]['Balance'].sum() +
            df[df['Affectation rapport '] == 'Fournisseurs et comptes rattachés']['Balance'].sum() +
            df[df['Affectation rapport '] == 'Concours bancaires et autres passifs financiers']['Balance'].sum()
        ) * -1

        total_liabilities = total_non_current_liabilities + total_current_liabilities

        total_equity = (
            -1 * df[df['Affectation rapport '] == 'Capital libéré']['Balance'].sum() +
            -1 * df[df['Affectation rapport '] == 'Reserve Legal']['Balance'].sum() +
            -1 * df[df['Affectation rapport '] == 'Résultat reporté']['Balance'].sum() +
            (operating_revenues - total_operating_expenses + other_gains_and_losses)
        )

        total_assets = (
            df[df['Classification'] == 'Assets']['Balance'].sum()
        )

        current_assets_total = (
            df[df['Affectation rapport '] == 'Stocks']['Balance'].sum() +
            df[df['Affectation rapport '] == 'Clients et comptes rattachés']['Balance'].sum() +
            df[df['Affectation rapport '] == 'Autres actifs courants']['Balance'].sum() +
            df[df['Affectation rapport '] == 'Liquidités et équivalents de liquidités']['Balance'].sum()
        )

        gross_profit = operating_revenues - total_operating_expenses
        net_income = gross_profit + other_gains_and_losses

        gross_profit_margin = (gross_profit / operating_revenues) if operating_revenues != 0 else None
        return_on_assets = (net_income / total_assets) if total_assets != 0 else None
        return_on_equity = (net_income / total_equity) if total_equity != 0 else None
        debt_ratio = total_liabilities / total_assets if total_assets != 0 else None
        equity_ratio = total_equity / total_assets if total_assets != 0 else None

        return pd.Series({
            'Gross Profit Margin': gross_profit_margin,
            'Return on Assets (ROA)': return_on_assets,
            'Return on Equity (ROE)': return_on_equity,
            'Debt Ratio': debt_ratio,
            'Equity Ratio': equity_ratio
        })

    # Calculate the financial ratios over time
    ratios_over_time = df_clean.groupby('Year').apply(calculate_financial_ratios).reset_index()

    # Plot the financial ratios over time
    fig_ratios = px.line(
        ratios_over_time, 
        x='Year', 
        y=['Return on Equity (ROE)', 'Return on Assets (ROA)', 'Debt Ratio', 'Equity Ratio', 'Gross Profit Margin'], 
        title='Trend Analysis of Financial Ratios Over Time',
        template='plotly_white'
    )
    fig_ratios.update_layout(
        xaxis_title='Year',
        yaxis_title='Ratio',
        legend_title='Ratios',
        font=dict(
            family="Arial, sans-serif",
            size=12,
            color="RebeccaPurple"
        )
    )
    st.plotly_chart(fig_ratios, use_container_width=True)


    """Display the AI assistant page content."""
    st.title("AI Assistant")
    st.write("Get assistance from our AI-powered financial assistant.")
    
    if 'chat_history' not in st.session_state:
        st.session_state.chat_history = []

    # Handle user input
    user_input = st.chat_input("You: ")

    if user_input:
        response, last_response = chatbot_response(user_input, st.session_state.last_response)
        st.session_state.last_response = last_response
        # Update chat history
        st.session_state.chat_history.append({"role": "user", "content": user_input})
        st.session_state.chat_history.append({"role": "assistant", "content": response})
    # Display chat history
    for message in st.session_state.chat_history: 
        with st.chat_message(message["role"]):
            st.markdown(message["content"])
         


def display_ai_assistant_page() -> None:
    """Display the AI assistant page content."""
    st.title("AI Assistant")
    st.write("Get assistance from our AI-powered financial assistant.")
    
    

def main() -> None:
    """Main function to run the Streamlit app."""
    st.set_page_config(
        page_title=PAGE_TITLE,
        page_icon=PAGE_ICON,
        layout="wide",
        initial_sidebar_state="expanded",
    )

    st.sidebar.title("Navigation")
    page = st.sidebar.radio("Go to", ["Welcome", "Upload & Process Excel", "KPIS", "Trend Analysis","AI Assistant"])

    if page == "Welcome":
        display_welcome_page()
    elif page == "Upload & Process Excel":
        lottie_upload = load_lottie_url(LOTTIE_URL)
        display_upload_page(lottie_upload)
    elif page == "KPIS":
        display_dashboards_page()
    #elif page == "AI Assistant":
        #display_ai_assistant_page()
    elif page == "Trend Analysis":
        display_trend_analysis_page()
    elif page == "AI Assistant":
        
          display_ai_assistant_page()
        
            
main()
